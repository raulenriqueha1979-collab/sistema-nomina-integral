"""Generación del informe en Excel (openpyxl): resumen + memoria de cálculo."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..calc.models import ResultadoLiquidacion

AZUL = "1F3864"
GRIS = "F2F2F2"
BLANCO = "FFFFFF"
FMT_BS = '"Bs. "#,##0.00'

_thin = Side(style="thin", color="BFBFBF")
BORDE = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _titulo(ws, celda, texto):
    ws[celda] = texto
    ws[celda].font = Font(bold=True, size=13, color=AZUL)


def _header_fill(cell):
    cell.fill = PatternFill("solid", fgColor=AZUL)
    cell.font = Font(bold=True, color=BLANCO)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDE


def generar_excel(resultado: ResultadoLiquidacion) -> bytes:
    """Genera el libro Excel (.xlsx) y devuelve los bytes."""
    wb = Workbook()

    # ================= Hoja 1: Resumen ejecutivo =================
    ws = wb.active
    ws.title = "Resumen"
    trab = resultado.trabajador
    par = resultado.parametros

    _titulo(ws, "A1", par.empresa_nombre or "Sistema de Control Fiscal RR")
    ws["A2"] = f"RIF: {par.empresa_rif}" if par.empresa_rif else ""
    _titulo(ws, "A3", "Recibo de Liquidación de Prestaciones Sociales (LOTTT)")

    info = [
        ("Trabajador", trab.nombre),
        ("Cédula de identidad", trab.cedula),
        ("Cargo", trab.cargo or "-"),
        ("Departamento", trab.departamento or "-"),
        ("Fecha de ingreso", str(trab.fecha_ingreso)),
        ("Fecha de egreso", str(trab.fecha_egreso)),
        ("Motivo de retiro", trab.motivo_retiro.etiqueta),
        ("Antigüedad", resultado.antiguedad.descripcion()),
    ]
    row = 5
    for k, v in info:
        ws.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws.cell(row=row, column=2, value=v)
        row += 1

    row += 1
    headers = ["Concepto", "Artículo", "Días", "Base (Bs.)", "Monto (Bs.)"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        _header_fill(cell)
    row += 1

    def _linea(concepto, articulo, dias, base, monto, negativo=False):
        nonlocal row
        ws.cell(row=row, column=1, value=("(-) " if negativo else "") + concepto).border = BORDE
        ws.cell(row=row, column=2, value=articulo).border = BORDE
        cd = ws.cell(row=row, column=3, value=float(dias) if dias is not None else None)
        cd.border = BORDE
        cb = ws.cell(row=row, column=4, value=float(base) if base is not None else None)
        cb.number_format = FMT_BS
        cb.border = BORDE
        cm = ws.cell(row=row, column=5, value=float(monto))
        cm.number_format = FMT_BS
        cm.border = BORDE
        row += 1

    for a in resultado.asignaciones:
        _linea(a.concepto, a.articulo, a.dias, a.base, a.monto)

    def _total(label, valor, color=GRIS, blanco=False):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(bold=True, color=BLANCO if blanco else "000000")
            cell.border = BORDE
        tot = ws.cell(row=row, column=5, value=float(valor))
        tot.number_format = FMT_BS
        tot.font = Font(bold=True, color=BLANCO if blanco else "000000")
        row += 1

    _total("TOTAL ASIGNACIONES", resultado.total_asignaciones)
    for d in resultado.deducciones:
        _linea(d.concepto, d.articulo, None, None, d.monto, negativo=True)
    _total("TOTAL DEDUCCIONES", resultado.total_deducciones)
    _total("NETO A PAGAR", resultado.neto_pagar, color=AZUL, blanco=True)

    widths = [42, 16, 10, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ================= Hoja 2: Memoria de cálculo =================
    ws2 = wb.create_sheet("Memoria de Cálculo")
    _titulo(ws2, "A1", "Memoria de Cálculo detallada (paso a paso)")
    ws2["A2"] = "Bases y fórmulas aplicadas conforme a la LOTTT"
    ws2["A2"].font = Font(italic=True, color="808080")

    bases = [
        ("Salario mensual normal", resultado.salario_mensual_normal),
        ("Salario diario normal (mensual / 30)", resultado.salario_diario_normal),
        ("Alícuota de utilidades", resultado.alicuota_utilidades),
        ("Alícuota de bono vacacional", resultado.alicuota_bono_vacacional),
        ("Salario integral diario", resultado.salario_integral_diario),
    ]
    r = 4
    ws2.cell(row=r, column=1, value="Base de cálculo").font = Font(bold=True)
    ws2.cell(row=r, column=2, value="Valor (Bs.)").font = Font(bold=True)
    r += 1
    for k, v in bases:
        ws2.cell(row=r, column=1, value=k)
        c = ws2.cell(row=r, column=2, value=float(v))
        c.number_format = FMT_BS
        r += 1

    r += 1
    ws2.cell(row=r, column=1, value="Detalle paso a paso").font = Font(bold=True, color=AZUL)
    r += 1
    for linea in resultado.memoria:
        ws2.cell(row=r, column=1, value="• " + linea)
        ws2.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    r += 1
    comparativa = [
        ("Días de garantía (Art. 142 a/b)", float(resultado.dias_garantia)),
        ("Días adicionales (Art. 142 c)", float(resultado.dias_adicionales)),
        ("Monto garantía acumulada", float(resultado.monto_garantia)),
        ("Días retroactivo (Art. 142 d)", float(resultado.dias_retroactivo)),
        ("Monto retroactivo", float(resultado.monto_retroactivo)),
        ("Método aplicado (monto mayor)", resultado.metodo_prestaciones),
        ("Prestaciones sociales adjudicadas", float(resultado.prestaciones_sociales)),
        ("Intereses (Art. 143)", float(resultado.intereses_prestaciones)),
        ("Indemnización (Art. 92)", float(resultado.indemnizacion)),
    ]
    ws2.cell(row=r, column=1, value="Comparativa de prestaciones (Art. 142)").font = Font(bold=True, color=AZUL)
    r += 1
    for k, v in comparativa:
        ws2.cell(row=r, column=1, value=k)
        cell = ws2.cell(row=r, column=2, value=v)
        if isinstance(v, float):
            cell.number_format = FMT_BS if v > 100 else "#,##0.00"
        r += 1

    ws2.column_dimensions["A"].width = 70
    ws2.column_dimensions["B"].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
